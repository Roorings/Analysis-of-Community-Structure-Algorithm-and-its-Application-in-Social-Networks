# import pacakges
import networkx as nx
from infomap import Infomap
from networkx.generators.community import LFR_benchmark_graph  
import pandas as pd                  
import igraph as ig                
from igraph import*
from openpyxl import Workbook
from networkx.algorithms.community import greedy_modularity_communities
import numpy as np
import sys

# define membership
# Copyright: https://github.com/Lonelycat1/Community-nullmodel
def network_community(community_list,G):    
    community_list_new = []
    for i in range(len(community_list)):
        temp = []
        nodes = community_list[i]
        for j in nodes:
            node = int(j-1)
            temp.append(node)
        community_list_new.append(temp)
    membership = []
    for k in range(0,len(G)):
       membership.append(0)
    for l in range(0,len(community_list_new)):
        nodes = community_list_new[l]
        for m in nodes:
            membership[m] = l
    for s in range(len(membership)):
        membership[s] = membership[s]+1
    return membership

#create excel to store the output
workbook = Workbook()
save_file = "/Users/yilingluo/Desktop/LFR_Benchmark.xlsx"
worksheet = workbook.active
w1 = workbook.create_sheet("fastgreedy") 
w2 = workbook.create_sheet("infomap")

#benchmark generation
nmifg = []
nmiinfomap = []
for i in range(1,9):
    G = LFR_benchmark_graph(250, 3, 1.5, mu, average_degree=5,min_community=20)

    # Fromat Transfomation & Membership Assignment
    communitiesLFR = {frozenset(G.nodes[v]["community"]) for v in G}
    lstLFR = [list(x) for x in communitiesLFR]
    membershipLFR = network_community(lstLFR,G)
    communitiesFG = greedy_modularity_communities(G, weight=None, resolution=1, n_communities=1)
    lstFG = [list(x) for x in communitiesFG]
    membershipFG = network_community(lstFG,G)
    origin_edges = nx.to_edgelist(G)
    new_list = []
    for edge_i in range(0,len(origin_edges)):
        node_dict = list(origin_edges)[edge_i]
        new_list.append([node_dict[0],node_dict[1]])
    g = Graph(new_list)  
    h1 = g.community_infomap()
    community_list_new = list(h1)
    list(community_list_new).remove(community_list_new[0])  
    membershipINFOMAP = network_community(community_list_new,G)

    nmifg.append(compare_communities(membershipFG, membershipLFR, method='nmi', remove_none=False))
    nmiinfomap.append(compare_communities(membershipINFOMAP, membershipLFR, method='nmi', remove_none=False))

#Append into Excel
w1.append(nmifg)
w2.append(nmiinfomap)
workbook.save(filename=save_file)

# GN Benchmark
lstin = []
lstfg = []
for j in range(1,15):
    pin = 0.38 - 0.02 * j
    pout = 4 / 600 + 4 / 600 * j
    avein = 0
    avefg = 0
    for k in range(20):
        G = nx.random_partition_graph([50, 50, 50, 50], pin, pout, directed = False)
        # Fromat Transfomation & Membership Assignment(omit)
        communitiesFG = greedy_modularity_communities(G, weight=None, resolution=1, n_communities=1)
        # Fromat Transfomation & Membership Assignment(omit)
        h1 = g.community_infomap()
        # Fromat Transfomation & Membership Assignment(omit)
        avein = avein + compare_communities(membershipINFOMAP, membershipRP, method='nmi', remove_none=False)
        avefg = avefg + compare_communities(membershipFG, membershipRP, method='nmi', remove_none=False)
    lstfg.append(avefg/20)
    lstin.append(avein/20)

# Construct a smaller DBLF Network
result = []
with open('com-dblp.top5000.cmty.txt','r') as f:
    for line in f:
        result.append(list(line.strip('\n').split(',')))
# Remove Duplication
set5000 = []
for i in range(len(result)):
    tryy = ''.join(result[i])
    a = set(tryy.split())
    set5000.append(a)
sort5000 = sorted(set5000, key = len, reverse = True)
for i in range(5000):
    for j in range(i+1,5000):
        sort5000[j] = sort5000[j] - sort5000[i]
final5000 = sorted(sort5000, key = len, reverse = True)
# Choose subgraph node
lstsub = []
for i in range(35,76):
    for every in iter(final5000[i]):
        lstsub.append(int(every))
G = nx.read_edgelist('com-dblp.ungraph.txt', comments='#', create_using=nx.Graph(),  delimiter=' ', nodetype=int, encoding='utf-8')
H = G.subgraph(lstsub)
# Remove no link nodes
remove = [node for node,degree in H.degree() if degree < 1]
F = nx.Graph(H)
F.remove_nodes_from(remove)

# Map nodes
# Copyright: https://blog.csdn.net/baidu_37995814/article/details/89710757
l1 = sorted(list(map(int,F.nodes())))
l2 = range(F.number_of_nodes())
nodes = dict(map(lambda x,y:[x,y],l1,l2)) 
edge_list=[]
for u,v in H.edges():
    edge_list.append((nodes[int(u)],nodes[int(v)]))
new_H=nx.Graph()
new_H.add_edges_from(edge_list)
nx.write_edgelist(new_H,'network_lianxu.txt',data=False)
orders = list(nodes)

# Map Membership
membershipGN = []
for k in range(len(orders)):
    for i in range(35,76):
        for j in range(len(final5000[i])):
            if(orders[k] == int(list(final5000[i])[j])):
                membershipGN.append(i-35)
membershipGNDF = pd.DataFrame(membershipGN)
membershipGNDF.to_csv('membershipGN.csv',index=False)

# Real Network Calculation
G = nx.read_edgelist('email-Eu-core.txt', comments='#', create_using=nx.Graph(), delimiter=' ',  nodetype=int, encoding='utf-8')
txt = np.loadtxt('email-Eu-core-department-labels.txt')
membershipGT = []
for i in range(len(txt)):
    membershipGT.append(txt[i][1])

communitiesFG = greedy_modularity_communities(G, weight=None, resolution=1, n_communities=1)
lstFG = [list(x) for x in communitiesFG]
membershipFG = network_community(lstFG,G)

imfb = Infomap("-k --include-self-links --clu --ftree -d --directed")
imfb.read_file("email-Eu-core.txt")
imfb.run()
membershipIN = []
for i in range(len(imfb.get_modules())):
    membershipIN.append(imfb.get_modules()[i])
print(compare_communities(membershipIN, membershipGT, method='nmi', remove_none=False))
print(compare_communities(membershipFG, membershipGT, method='nmi', remove_none=False))

# Random Geometric Graph & Star Graph
# Copyright: https://zhuanlan.zhihu.com/p/348070109
G = nx.random_geometric_graph(150, 0.125)
pos = nx.get_node_attributes(G, "pos")
dmin = 1
ncenter = 0
for n in pos:
    x, y = pos[n]
    d = (x - 0.5) ** 2 + (y - 0.5) ** 2
    if d < dmin:
        ncenter = n
        dmin = d
p = dict(nx.single_source_shortest_path_length(G, ncenter))
plt.figure(figsize=(8, 8))
nx.draw_networkx_edges(G, pos, nodelist=[ncenter], alpha=0.4)
nx.draw_networkx_nodes(G, pos, nodelist=list(p.keys()), node_size=80, node_color=list(p.values()), cmap=plt.cm.Reds_r,)
plt.xlim(-0.05, 1.05)
plt.ylim(-0.05, 1.05)
plt.axis("off")
plt.show()
G = nx.star_graph(30)
pos = nx.spring_layout(G)
colors = range(30)
nx.draw(G, pos, node_color='#A0CBE2', edge_color=colors,width=2, edge_cmap=plt.cm.Blues, with_labels=False)
plt.show()

# Find the intersection of  the partition
# Do the same for others(omit)
c =[0] * max(membershipIN)
list_IN = [[0] for j in range(int(max(membershipIN)))]
for i in range(len(membershipIN)):
    c[int(membershipIN[i])-1] = c[int(membershipIN[i])-1]+1
    list_IN[int(membershipIN[i])-1].append(i)
print(c)
for i in range(int(max(membershipIN))):
    list_IN[i] = set(list_IN[i])
list_IN_sort = sorted(list_IN, key = len, reverse = True)
print(list_IN_sort)
for j in range(0,10):
    maxx = 0
    print('a')
    for i in range(len(list_FG)):
        if (1 - len(list_GT_sort[j] - list_FG_sort[i]) / len(list_GT_sort[j])) > maxx:
            maxx = 1 - len(list_GT_sort[j] - list_FG_sort[i]) / len(list_GT_sort[j])
            print(maxx)
    print(maxx)
