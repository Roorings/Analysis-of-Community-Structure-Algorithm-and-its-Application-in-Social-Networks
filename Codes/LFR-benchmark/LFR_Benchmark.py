# import pacakges
import networkx as nx
from infomap import Infomap
from networkx.generators.community import LFR_benchmark_graph  
import pandas as pd                  
import igraph as ig                
from igraph import*
from openpyxl import Workbook
from networkx.algorithms.community import greedy_modularity_communities
import numpy as np

# define membership
def network_community(community_list,G):    
    community_list_new = []
    for i in range(len(community_list)):
        temp = []
        nodes = community_list[i]
        for j in nodes:
            node = int(j-1)
            temp.append(node)
        community_list_new.append(temp)
        
    membership = []
    for k in range(0,len(G)):
       membership.append(0)

    for l in range(0,len(community_list_new)):
        nodes = community_list_new[l]
        for m in nodes:
            membership[m] = l
    for s in range(len(membership)):
        membership[s] = membership[s]+1
        
    return membership

#create excel to store the output
workbook = Workbook()
save_file = "/Users/yilingluo/Desktop/LFR_Benchmark.xlsx"
worksheet = workbook.active
w1 = workbook.create_sheet("fastgreedy") 
w2 = workbook.create_sheet("infomap")

#benchmark generation
nmifg = []
nmiinfomap = []
for i in range(2,17):
    n = 250
    tau1 = 3
    tau2 = 1.5
    mu = i/20
    G = LFR_benchmark_graph(n, tau1, tau2, mu, average_degree=5,min_community=20)
    print(mu)

    communitiesLFR = {frozenset(G.nodes[v]["community"]) for v in G}
    lstLFR = [list(x) for x in communitiesLFR]
    membershipLFR = network_community(lstLFR,G)

    communitiesFG = greedy_modularity_communities(G, weight=None, resolution=1, n_communities=1)
    lstFG = [list(x) for x in communitiesFG]
    membershipFG = network_community(lstFG,G)

    origin_edges = nx.to_edgelist(G)
    new_list = []
    for edge_i in range(0,len(origin_edges)):
        node_dict = list(origin_edges)[edge_i]
        new_list.append([node_dict[0],node_dict[1]])
    g = Graph(new_list)  
    h1 = g.community_infomap()
    community_list_new = list(h1)
    list(community_list_new).remove(community_list_new[0])  
    membershipINFOMAP = network_community(community_list_new,G)

    print(compare_communities(membershipFG, membershipLFR, method='nmi', remove_none=False))
    nmifg.append(compare_communities(membershipFG, membershipLFR, method='nmi', remove_none=False))
    nmiinfomap.append(compare_communities(membershipINFOMAP, membershipLFR, method='nmi', remove_none=False))

w1.append(nmifg)
w2.append(nmiinfomap)
workbook.save(filename=save_file)

#typical benchmark graph
n = 250
tau1 = 3
tau2 = 1.5
mu = 0.5
G = LFR_benchmark_graph(n, tau1, tau2, mu, average_degree=5,min_community=20,seed = 10)
nx.draw(G, with_labels=True)
#LFR benchmark
c = {frozenset(G.nodes[v]["community"]) for v in G}
#FN algorithm
from networkx.algorithms.community import greedy_modularity_communities 
c = greedy_modularity_communities(G, weight=None, resolution=1, n_communities=1)
partition_map = {}
for idx, cluster_nodes in enumerate(c):
    for node in cluster_nodes:
        partition_map[node] = idx
node_colors = [partition_map[n] for n in G.nodes]       
#Infomap algorithm
nx.write_edgelist(G, "test.txt") # using Infomap 1.0.x
imfb = Infomap("--clu --ftree")
imfb.read_file("test.txt")
imfb.run()
tree = imfb
communities = {}
for node in tree.iterTree():
    communities[node.physicalId] = node.moduleIndex()
nx.set_node_attributes(G, name='community', values=communities)
node_colors = [communities[n] for n in G.nodes]
#print a graph
nx.draw(G, node_color=node_colors, with_labels=True)